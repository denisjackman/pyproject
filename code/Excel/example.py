#!/usr/bin/env python
'''
    excel worksheet work
    https://realpython.com/openpyxl-excel-spreadsheets-python/
'''
import datetime
from openpyxl import Workbook


def main():
    '''
        main function
    '''
    wb = Workbook()
    ws1 = wb.create_sheet("Denis Sheet")
    ws2 = wb.create_sheet("Other Sheet")

    ws1['A1'] = 42
    ws1['C20'] = 'hello world'
    ws = wb.active

    for column in "BCDE":
        for row in range(2, 11):
            target = column+str(row)
            ws[target] = 'X'

    ws2.append([1, 2, 3])

    # Python types will automatically be convertedpy merge
    ws2['A2'] = datetime.datetime.now()

    # Save the file
    wb.save("y:/Resources/excel/Denis.xlsx")


if __name__ == "__main__":
    main()
