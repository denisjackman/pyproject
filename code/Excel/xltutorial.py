'''
    https://www.youtube.com/watch?v=7YS6YDQKFh0

OpenPyXL Docs: https://openpyxl.readthedocs.io/en/st...
Code Written in This Tutorial: https://github.com/techwithtim/ExcelP...
Fix Pip (Windows): https://www.youtube.com/watch?v=AdUZA...
Fix Pip (Mac/Linux): https://www.youtube.com/watch?v=E-WhA...

works with excel from version 2010 onwards
'''
from openpyxl import load_workbook


def main():
    ''' main function '''
    wb = load_workbook('y:/Resources/excel/grades.xlsx')
    ws = wb.active
    ws['A10'].value = 'DemonBoy666'

    print(wb.sheetnames)
    print(wb['Sheet1'])
    wb.create_sheet('Test')
    print(wb.sheetnames)

    wb.save('y:/Resources/excel/grades1.xlsx')
    ws = wb['Grades']
    print(ws['A2'].value)
    print(ws['Z99'].value is None)


if __name__ == '__main__':
    main()
