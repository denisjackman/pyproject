'''
a contining series of xl tutorials
'''
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def main():
    '''	main function	'''
    data = {"Joe": {"math": 65, "science": 78, "english": 98, "gym": 89},
            "Bill": {"math": 55, "science": 72, "english": 87, "gym": 95},
            "Tim": {"math": 100, "science": 45, "english": 75, "gym": 92},
            "Sally": {"math": 30, "science": 25, "english": 45, "gym": 100},
            "Jane": {"math": 100, "science": 100, "english": 100, "gym": 60}}

    wb = Workbook()
    ws = wb.active
    ws.title = "New Grades"
    headings = ['Name'] + list(data['Joe'].keys())
    ws.append(headings)
    for person, grades in data.items():
        print(f'person is {person} and grades is {grades}')
        tempstr = [person] + list(grades.values())
        ws.append(tempstr)

    for col in range(2, len(data['Joe']) + 2):
        char = get_column_letter(col)
        ws[char + '7'] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

    for newcol in range(1, 6):
        cell = ws[get_column_letter(newcol) + '1']
        cell.font = Font(bold=True)

    wb.save('y:/Resources/excel/NewGrades2.xls')


if __name__ == '__main__':
    main()
