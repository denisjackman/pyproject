'''
a contining series of xl tutorials
'''
from openpyxl import load_workbook


def main():
    ''' main function '''
    wb = load_workbook('y:/Resources/excel/tim.xlsx')
    ws = wb.active

    ws.merge_cells('A1:E1')
    ws.merge_cells('A2:E2')

    ws.unmerge_cells('A1:E1')

    wb.save('y:/Resources/excel/tim2.xlsx')


if __name__ == '__main__':
    main()
