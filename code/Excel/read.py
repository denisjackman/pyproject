'''
 Reading an excel file using Python
'''
import openpyxl


def main():
    ''' main function   '''
    # Give the location of the file
    loc = openpyxl.load_workbook("y:/Resources/excel/test.xlsx")

    sheet = loc.active

    # For row 2 and column 1
    print(sheet.cell(row=2, column=1).value)
    print(sheet.cell(row=3, column=1).value)
    print(sheet.cell(row=4, column=1).value)


if __name__ == '__main__':
    main()
