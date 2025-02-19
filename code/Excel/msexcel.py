"""
msexcel.py

This program is a template for python programs
All this stuff at the top of the script is just optional metadata;

"""

__author__ = "Denis J Jackman (denis_jackman@hotmail.com)"
__version__ = "$Revision: 1.10 $"
__date__ = "$Date: 2022/06/08 16:47:00 $"
__copyright__ = "Copyright (c) 2022 Denis J Jackman"
__license__ = "Python"

from openpyxl import Workbook
from openpyxl import load_workbook


def main():
    ''' main function '''
    wb = Workbook()
    ws = wb.active
    wb.create_sheet("FirstSheet")
    wb.create_sheet("SecondSheet", 0)
    wb.create_sheet("ThirdSheet", -1)
    wb.save('y:/Resources/excel/openpyxl_example.xlsx')

    ws.title = "OpenPyXL Example"
    newwb = load_workbook('y:/Resources/excel/example.xlsx')
    print(type(newwb))


if __name__ == '__main__':
    main()
