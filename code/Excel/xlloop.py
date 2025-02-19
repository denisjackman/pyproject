'''
a contining series of xl tutorials
'''
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def main():
    ''' main function'''
    wb = load_workbook('y:/Resources/excel/tim.xlsx')
    ws = wb.active
    for xlrow in range(1, 11):
        for xlcol in range(1, 5):
            char = get_column_letter(xlcol)
            print(ws[char + str(xlrow)].value)
            ws[char + str(xlrow)] = char + str(xlrow)

    ws.append(['Tim', 'Is', 'Great', 'Always', '!'])
    ws.append(['Tim', 'Is', 'Great', 'Always', '!'])
    ws.append(['Tim', 'Is', 'Great', 'Always', '!'])
    ws.append(['Tim', 'Is', 'Great', 'Always', '!'])
    ws.append(['end'])
    wb.save('y:/Resources/excel/tim4.xlsx')
    print(f'Your version is {sys.version_info[0]}.{sys.version_info[1]} ')


if __name__ == '__main__':
    main()
